import openpyxl

file = "C:/Users/com/Downloads/Login.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active     #or sheet = workbook['Data'] or sheet = workbook['Sheet1']

# to write same value everywhere
# for r in range(1,4):
#     for c in range(1,4):
#         sheet.cell(r,c).value='welcome'
# workbook.save(file)

# to write different values

sheet.cell(1,1).value = 'EMP ID'
sheet.cell(1,2).value = 'Name'

sheet.cell(2,1).value = '12'
sheet.cell(2,2).value = 'Xavier'

sheet.cell(3,1).value = '34'
sheet.cell(3,2).value = 'Thapar'

sheet.cell(4,1).value = '56'
sheet.cell(4,2).value = 'Nehru'

workbook.save(file)
