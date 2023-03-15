import openpyxl

#----> file---> workbook---> sheet---> rows---> columns

file = "C:/Users/com/Downloads/LoginData.xlsx"      # path of the file
workbook = openpyxl.load_workbook(file)     #workbook access
sheet = workbook["Sheet1"]          #Sheet access

rows = sheet.max_row            #counts the number of rows in excel
columns = sheet.max_column      #counts the number of columns in the excel

# prints the values present in the excel sheet
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end='  ')
    print()
